# prism/backend/app/utils/file_parser.py
"""文件解析器：将不同格式文件提取为纯文本。"""
import os
from pathlib import Path


def extract_text(file_path: str) -> str:
    """根据扩展名分发到对应解析器，返回纯文本。"""
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return _extract_pdf(file_path)
    if ext == ".docx":
        return _extract_docx(file_path)
    if ext == ".xlsx":
        return _extract_xlsx(file_path)
    if ext in (".md", ".txt", ".markdown"):
        return Path(file_path).read_text(encoding="utf-8")
    raise ValueError(f"不支持的文件类型: {ext}")


def _extract_pdf(file_path: str) -> str:
    import fitz  # PyMuPDF
    doc = fitz.open(file_path)
    text_parts = []
    for page in doc:
        text_parts.append(page.get_text())
    doc.close()
    return "\n".join(text_parts)


def _extract_docx(file_path: str) -> str:
    from docx import Document
    doc = Document(file_path)
    return "\n".join(para.text for para in doc.paragraphs)


def _extract_xlsx(file_path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    text_parts = []
    for sheet in wb.worksheets:
        text_parts.append(f"## 工作表: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                text_parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(text_parts)


def extract_url(url: str) -> str:
    """抓取网页，提取正文。"""
    import httpx
    import re
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers={"User-Agent": "Prism/1.0"})
    resp.raise_for_status()
    html = resp.text
    # 简易正文提取：去标签
    text = re.sub(r"<script[^>]*>.*?</script>", "", html, flags=re.DOTALL)
    text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL)
    text = re.sub(r"<[^>]+>", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()
